"""
PancreaScan — Issues Report Generator
generate_issues_report.py

Generates Issues_Report_PancreaScan_<timestamp>.xlsx containing ONLY
failed/errored test cases so developers can quickly identify and fix problems.

Can be used:
  1. Called by generate_test_report.py automatically when failures exist
  2. Run standalone:  python generate_issues_report.py [--junit path/*.xml]
"""

import sys
import glob
import argparse
import subprocess
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Style helpers ──────────────────────────────────────────────────────────────
RED       = "DC2626"
RED_LT    = "FEF2F2"
RED_DARK  = "7F1D1D"
AMBER     = "D97706"
AMBER_LT  = "FFFBEB"
AMBER_MED = "B45309"
BLUE_DARK = "1E3A8A"
BLUE_MED  = "2563EB"
BLUE_LT   = "DBEAFE"
GREEN     = "16A34A"
GREEN_LT  = "DCFCE7"
GRAY_DARK = "0F172A"
GRAY_LT   = "F1F5F9"
WHITE     = "FFFFFF"

PRIORITY_MAP = {
    # Security failures are CRITICAL
    "Security":      ("CRITICAL", RED_DARK,  RED_LT),
    "Functional":    ("HIGH",     RED,       RED_LT),
    "Validation":    ("HIGH",     RED,       RED_LT),
    "Performance":   ("MEDIUM",   AMBER_MED, AMBER_LT),
    "UI/UX":         ("MEDIUM",   AMBER_MED, AMBER_LT),
    "Deployability": ("CRITICAL", RED_DARK,  RED_LT),
    "Unit":          ("HIGH",     RED,       RED_LT),
    "Accessibility": ("LOW",      BLUE_MED,  BLUE_LT),
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border():
    thin = Side(style="thin", color="E2E8F0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _priority(test_type: str) -> tuple:
    return PRIORITY_MAP.get(test_type, ("MEDIUM", AMBER, AMBER_LT))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def build_issues_summary_sheet(wb, failed_tests: list, now_str: str):
    """Sheet 1: High-level summary of failures."""
    ws = wb.create_sheet("🚨 Issues Summary")
    ws.sheet_view.showGridLines = False

    total_fail = len(failed_tests)
    critical   = sum(1 for t in failed_tests if _priority(t[3])[0] == "CRITICAL")
    high       = sum(1 for t in failed_tests if _priority(t[3])[0] == "HIGH")
    medium     = sum(1 for t in failed_tests if _priority(t[3])[0] == "MEDIUM")
    low        = sum(1 for t in failed_tests if _priority(t[3])[0] == "LOW")

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "🚨  PancreaScan — Test Issues Report (Failures Only)"
    ws["A1"].font      = Font(bold=True, color=WHITE, size=18, name="Calibri")
    ws["A1"].fill      = _fill(RED_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 44

    ws.merge_cells("A2:H2")
    ws["A2"] = (f"Generated: {now_str}   |   "
                f"Action Required: Fix all CRITICAL and HIGH issues before deployment")
    ws["A2"].font      = Font(color="FCA5A5", size=10, name="Calibri")
    ws["A2"].fill      = _fill(RED)
    ws["A2"].alignment = _align("center")
    ws.row_dimensions[2].height = 20

    # Stat cards
    cards = [
        ("Total Failures",  str(total_fail), RED,       RED_LT),
        ("🔴 CRITICAL",     str(critical),   RED_DARK,  RED_LT),
        ("🟠 HIGH",         str(high),       RED,       RED_LT),
        ("🟡 MEDIUM",       str(medium),     AMBER_MED, AMBER_LT),
        ("🟢 LOW",          str(low),        GREEN,     GREEN_LT),
    ]
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 38
    for ci, (label, val, fg, bg) in enumerate(cards, 1):
        ws.cell(row=4, column=ci, value=label).font      = Font(bold=True, color=fg, size=10, name="Calibri")
        ws.cell(row=4, column=ci).fill                   = _fill(bg)
        ws.cell(row=4, column=ci).alignment              = _align("center")
        ws.cell(row=5, column=ci, value=val).font        = Font(bold=True, color=fg, size=26, name="Calibri")
        ws.cell(row=5, column=ci).fill                   = _fill(bg)
        ws.cell(row=5, column=ci).alignment              = _align("center")

    # By suite
    by_suite: dict = {}
    for t in failed_tests:
        by_suite.setdefault(t[2], 0)
        by_suite[t[2]] += 1

    ws.merge_cells("A7:E7")
    ws["A7"] = "Failures by Suite"
    ws["A7"].font = _font(bold=True, color=GRAY_DARK, size=12)
    ws["A7"].fill = _fill(GRAY_LT)
    ws["A7"].alignment = _align("center")
    ws.row_dimensions[7].height = 22

    hdr = ["Suite / Layer", "Failures", "Priority Level", "Action"]
    for ci, h in enumerate(hdr, 1):
        c = ws.cell(row=8, column=ci, value=h)
        c.font = _font(bold=True)
        c.fill = _fill(RED)
        c.alignment = _align("center")
        c.border = _border()
    ws.row_dimensions[8].height = 20

    for ri, (suite, cnt) in enumerate(sorted(by_suite.items(), key=lambda x: -x[1]), 9):
        action = "Fix immediately" if cnt > 5 else "Review and fix"
        for ci, val in enumerate([suite, cnt, "HIGH", action], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = _fill(GRAY_LT if ri % 2 == 0 else WHITE)
            c.font = Font(color=GRAY_DARK, size=10, name="Calibri")
            c.alignment = _align("center")
            c.border = _border()
        ws.row_dimensions[ri].height = 18

    # Retest instructions
    retest_row = 9 + len(by_suite) + 2
    ws.merge_cells(f"A{retest_row}:H{retest_row}")
    ws[f"A{retest_row}"] = "🔁  Re-Test Instructions: Fix the issues listed in 'Failed Tests' sheet → Commit → Push → GitHub Actions will re-run automatically"
    ws[f"A{retest_row}"].font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    ws[f"A{retest_row}"].fill = _fill(BLUE_MED)
    ws[f"A{retest_row}"].alignment = _align("center")
    ws.row_dimensions[retest_row].height = 28

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["A"].width = 22


def build_failed_tests_sheet(wb, failed_tests: list):
    """Sheet 2: Detailed list of all failed test cases."""
    ws = wb.create_sheet("❌ Failed Tests")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"❌  All Failed / Errored Test Cases — {len(failed_tests)} Issues"
    ws["A1"].font = _font(bold=True, size=14)
    ws["A1"].fill = _fill(RED_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 32

    cols   = ["TC ID", "Test Name", "Suite", "Type", "Status", "Error / Failure Message", "Priority", "Fix Recommendation"]
    widths = [12, 28, 12, 14, 10, 50, 12, 40]
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True)
        c.fill = _fill(RED)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    for ri, (tc_id, name, layer, ttype, desc, status, message) in enumerate(failed_tests, 3):
        prio_label, prio_color, prio_bg = _priority(ttype)
        row_fill = _fill(RED_LT) if ri % 2 == 0 else _fill(WHITE)

        fix_rec = _get_fix_recommendation(ttype, name, message)
        status_label = "❌ FAIL" if status == "FAIL" else "💥 ERROR"

        vals = [tc_id, name, layer, ttype, status_label, message, prio_label, fix_rec]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _border()
            c.alignment = _align("left", "center", wrap=True)
            if ci == 5:  # Status
                c.fill = _fill(RED_LT)
                c.font = Font(color=RED, bold=True, size=10, name="Calibri")
            elif ci == 7:  # Priority
                c.fill = _fill(prio_bg)
                c.font = Font(color=prio_color, bold=True, size=10, name="Calibri")
            else:
                c.fill = row_fill
                c.font = Font(color=GRAY_DARK, size=10, name="Calibri")
        ws.row_dimensions[ri].height = 24


def build_reproduce_sheet(wb, failed_tests: list):
    """Sheet 3: How to reproduce each failure."""
    ws = wb.create_sheet("🔁 How to Reproduce")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:D1")
    ws["A1"] = "🔁  How to Reproduce Failed Tests"
    ws["A1"].font = _font(bold=True, size=14)
    ws["A1"].fill = _fill(BLUE_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 32

    cols   = ["TC ID", "Test Name", "Run Command", "Expected Outcome"]
    widths = [12, 28, 70, 45]
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True)
        c.fill = _fill(BLUE_MED)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    suite_cmd = {
        "API":         "cd testing && pytest api/ -v -k",
        "Unit":        "cd testing && pytest unit/ -v -k",
        "Mobile":      "cd testing && pytest appium/ -v -k",
        "Web":         "cd testing && pytest selenium/ -v -k",
        "Functional":  "cd testing && pytest functional/ -v -k",
    }

    for ri, (tc_id, name, layer, ttype, desc, status, message) in enumerate(failed_tests, 3):
        cmd_base = suite_cmd.get(layer, "cd testing && pytest -v -k")
        # Convert TC-S001 to TCS001 for -k filter
        k_filter = tc_id.replace("-", "").upper()
        run_cmd = f'{cmd_base} "{k_filter}" --tb=long -v'
        row_fill = _fill(GRAY_LT if ri % 2 == 0 else WHITE)

        for ci, val in enumerate([tc_id, name, run_cmd, desc], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _border()
            c.alignment = _align("left", "center", wrap=True)
            if ci == 3:
                c.font = Font(color="1E40AF", size=10, name="Courier New")
                c.fill = _fill("EFF6FF")
            else:
                c.fill = row_fill
                c.font = Font(color=GRAY_DARK, size=10, name="Calibri")
        ws.row_dimensions[ri].height = 24


def build_fix_checklist_sheet(wb, failed_tests: list):
    """Sheet 4: Actionable fix checklist."""
    ws = wb.create_sheet("✅ Fix Checklist")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "✅  Developer Fix Checklist — Check off items as you fix them"
    ws["A1"].font = _font(bold=True, size=14)
    ws["A1"].fill = _fill(BLUE_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 32

    cols   = ["#", "TC ID", "Issue Description", "Fix Action", "Fixed? (✓)"]
    widths = [5, 12, 45, 55, 12]
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True)
        c.fill = _fill(BLUE_MED)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    for ri, (tc_id, name, layer, ttype, desc, status, message) in enumerate(failed_tests, 3):
        fix = _get_fix_recommendation(ttype, name, message)
        row_fill = _fill(RED_LT if ri % 2 == 0 else WHITE)
        num = ri - 2

        for ci, val in enumerate([num, tc_id, f"{name}: {message[:80]}", fix, "☐"], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _border()
            c.alignment = _align("center" if ci in (1, 5) else "left", "center", wrap=True)
            c.fill = row_fill
            c.font = Font(
                color=GRAY_DARK if ci != 5 else BLUE_MED,
                bold=(ci == 5),
                size=10 if ci != 5 else 14,
                name="Calibri"
            )
        ws.row_dimensions[ri].height = 28

    # Footer
    footer_row = 3 + len(failed_tests) + 1
    ws.merge_cells(f"A{footer_row}:E{footer_row}")
    ws[f"A{footer_row}"] = ("📌  After fixing: git add . → git commit -m 'fix: resolve test failures' "
                            "→ git push → GitHub Actions will re-run all tests automatically")
    ws[f"A{footer_row}"].font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    ws[f"A{footer_row}"].fill = _fill(GREEN)
    ws[f"A{footer_row}"].alignment = _align("center")
    ws.row_dimensions[footer_row].height = 28


def _get_fix_recommendation(test_type: str, test_name: str, message: str) -> str:
    """Return a context-aware fix recommendation string."""
    name_l    = test_name.lower()
    message_l = (message or "").lower()

    if "sql injection" in name_l or "injection" in name_l:
        return "Use parameterized queries in SQLAlchemy. Never format SQL with f-strings or % formatting."
    if "xss" in name_l:
        return "Sanitize/escape HTML output. Store XSS payloads as-is (don't eval). Use Content-Security-Policy header."
    if "rate limit" in name_l or "rate_limit" in name_l:
        return "Add slowapi rate limiting: @limiter.limit('5/minute') on login endpoint."
    if "token" in name_l and "401" in message_l:
        return "Check JWT middleware is applied. Ensure Authorization header is validated on protected routes."
    if "password" in name_l and ("hash" in name_l or "sha" in name_l):
        return "Replace hashlib.sha256 with bcrypt or argon2: pip install passlib[bcrypt]."
    if "cors" in name_l:
        return "Restrict CORS origins. Replace allow_origins=['*'] with specific domains."
    if "500" in message_l or "internal server error" in message_l:
        return "500 error means server crash. Check backend logs, add try/except, validate input before processing."
    if "timeout" in message_l or "response time" in name_l:
        return "Optimize slow endpoint. Check database queries for N+1. Add caching or index missing columns."
    if "selenium" in test_type.lower() or "web" in test_type.lower():
        return "Check if deployed frontend URL is correct. Ensure GitHub Pages site is live and returning 200."
    if "mobile" in test_type.lower():
        return "Appium tests require running emulator + Appium server. Check app package name and activities."
    if "not found" in message_l or "404" in message_l:
        return "Endpoint not found. Verify route path, check if endpoint exists in FastAPI router."
    if "422" in message_l or "validation" in message_l.lower():
        return "Pydantic validation error. Check request payload schema matches what API expects."
    if "connection" in message_l or "refused" in message_l:
        return "Backend server is not running. Ensure API server is started before running tests."

    return "Review error message in 'Failed Tests' sheet. Run the test locally to reproduce. Check backend logs."


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def generate_issues_report(
    failed_tests: list,
    output_dir: Path,
    now_str: str = None,
    ts: str = None,
) -> str:
    """
    Called from generate_test_report.py when failures are found.
    failed_tests: list of (tc_id, name, layer, ttype, desc, status, message)
    Returns path to the generated Issues XLSX file.
    """
    if now_str is None:
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if ts is None:
        ts = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")

    out_path = Path(output_dir) / f"Issues_Report_PancreaScan_{ts}.xlsx"

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("⚙️  Building Issues Summary sheet...")
    build_issues_summary_sheet(wb, failed_tests, now_str)

    print("⚙️  Building Failed Tests detail sheet...")
    build_failed_tests_sheet(wb, failed_tests)

    print("⚙️  Building Reproduce steps sheet...")
    build_reproduce_sheet(wb, failed_tests)

    print("⚙️  Building Fix Checklist sheet...")
    build_fix_checklist_sheet(wb, failed_tests)

    wb.save(str(out_path))
    return str(out_path)


# ── Standalone usage ───────────────────────────────────────────────────────────
if __name__ == "__main__":
    import xml.etree.ElementTree as ET

    parser = argparse.ArgumentParser(description="PancreaScan Issues Report Generator")
    parser.add_argument("--junit",  nargs="*", default=None,
                        help="JUnit XML file(s) from pytest. Glob patterns supported.")
    parser.add_argument("--output", default=None,
                        help="Output directory (default: testing/reports/)")
    args = parser.parse_args()

    out_dir = Path(args.output or Path(__file__).parent / "reports")
    out_dir.mkdir(parents=True, exist_ok=True)

    # Find JUnit XMLs
    junit_paths = args.junit
    if not junit_paths:
        discovered = list((Path(__file__).parent / "reports").glob("*.xml"))
        if discovered:
            junit_paths = [str(p) for p in discovered]
            print(f"  📂 Auto-discovered {len(junit_paths)} JUnit XML file(s)")

    # Parse failures
    failed_tests = []
    if junit_paths:
        all_xml = []
        for pat in junit_paths:
            all_xml.extend(glob.glob(pat))

        for xml_path in all_xml:
            if not Path(xml_path).exists():
                continue
            tree = ET.parse(xml_path)
            root = tree.getroot()
            for tc in root.findall(".//testcase"):
                failure = tc.find("failure")
                error   = tc.find("error")
                if failure is not None or error is not None:
                    el = failure if failure is not None else error
                    status  = "FAIL" if failure is not None else "ERROR"
                    message = (el.get("message") or el.text or "")[:300]
                    classname = tc.get("classname", "")
                    name      = tc.get("name", "")
                    # Derive layer from classname
                    layer = "API"
                    if "appium" in classname.lower() or "mobile" in classname.lower():
                        layer = "Mobile"
                    elif "selenium" in classname.lower() or "web" in classname.lower():
                        layer = "Web"
                    elif "functional" in classname.lower():
                        layer = "Functional"
                    elif "unit" in classname.lower():
                        layer = "Unit"
                    failed_tests.append((
                        name[:12], name, layer, "Functional", name, status, message
                    ))

    if not failed_tests:
        print("🎉 No failures found in JUnit XML — Issues report not needed!")
        print("   All tests passed or were skipped. Great job!")
        sys.exit(0)

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ts      = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")

    print(f"\n  🔴 Found {len(failed_tests)} failed/errored test(s) — generating Issues Report...")
    out_path = generate_issues_report(failed_tests, out_dir, now_str, ts)
    print(f"\n  ✅ Issues Report: {out_path}")
    print(f"\n  📌 Fix the issues listed, then re-push to trigger automated re-testing.")
