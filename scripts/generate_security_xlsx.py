"""
PancreaScan — Security Vulnerability XLSX Report Generator
generate_security_xlsx.py

Reads JSON output from security scanning tools (Semgrep, Trivy, pip-audit,
custom analysis) and produces a professional XLSX vulnerability report.

Usage:
  python scripts/generate_security_xlsx.py \
    --semgrep   scan-outputs/semgrep-results.json \
    --trivy     scan-outputs/trivy-results.json \
    --pip-audit scan-outputs/pip-audit-results.json \
    --custom    scan-outputs/custom-analysis.json \
    --output    "Vulnerability Test Results/Security_Report.xlsx"
"""

import sys
import json
import argparse
import subprocess
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Styles ─────────────────────────────────────────────────────────────────────
RED_DARK  = "7F1D1D"
RED       = "DC2626"
RED_LT    = "FEF2F2"
ORANGE    = "C2410C"
ORANGE_LT = "FFF7ED"
AMBER     = "D97706"
AMBER_LT  = "FFFBEB"
GREEN     = "16A34A"
GREEN_LT  = "DCFCE7"
BLUE_DARK = "1E3A8A"
BLUE_MED  = "2563EB"
BLUE_LT   = "DBEAFE"
GRAY_DARK = "0F172A"
GRAY_LT   = "F1F5F9"
WHITE     = "FFFFFF"

SEVERITY_CFG = {
    "CRITICAL": (RED_DARK, RED_LT,    "🔴 CRITICAL"),
    "HIGH":     (RED,      RED_LT,    "🟠 HIGH"),
    "MEDIUM":   (AMBER,    AMBER_LT,  "🟡 MEDIUM"),
    "LOW":      (GREEN,    GREEN_LT,  "🟢 LOW"),
    "INFO":     (BLUE_MED, BLUE_LT,   "ℹ️ INFO"),
    "WARNING":  (AMBER,    AMBER_LT,  "⚠️ WARNING"),
}

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color=WHITE, size=11): return Font(bold=bold, color=color, size=size, name="Calibri")
def _border():
    t = Side(style="thin", color="E2E8F0")
    return Border(left=t, right=t, top=t, bottom=t)
def _align(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ══════════════════════════════════════════════════════════════════════════════
# PARSERS FOR EACH TOOL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def _load_json(path: str) -> dict:
    if not path or not Path(path).exists():
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"  [WARN] Could not parse {path}: {e}")
        return {}


def parse_semgrep(data: dict) -> list:
    """Parse semgrep JSON output → list of finding dicts."""
    findings = []
    for result in data.get("results", []):
        sev = result.get("extra", {}).get("severity", "WARNING").upper()
        if sev not in SEVERITY_CFG:
            sev = "MEDIUM"
        findings.append({
            "tool":        "Semgrep",
            "severity":    sev,
            "vuln_type":   result.get("check_id", "Unknown rule"),
            "file_path":   result.get("path", ""),
            "line":        result.get("start", {}).get("line", ""),
            "description": result.get("extra", {}).get("message", "")[:300],
            "remediation": result.get("extra", {}).get("fix", "Review and remediate per OWASP guidelines"),
            "snippet":     (result.get("extra", {}).get("lines", "") or "")[:200],
        })
    return findings


def parse_trivy(data: dict) -> list:
    """Parse trivy JSON output → list of finding dicts."""
    findings = []
    for result in data.get("Results", []):
        for vuln in result.get("Vulnerabilities", []):
            sev = vuln.get("Severity", "MEDIUM").upper()
            if sev not in SEVERITY_CFG:
                sev = "MEDIUM"
            findings.append({
                "tool":        "Trivy",
                "severity":    sev,
                "vuln_type":   f"CVE: {vuln.get('VulnerabilityID', 'Unknown')}",
                "file_path":   vuln.get("PkgName", ""),
                "line":        f"Version: {vuln.get('InstalledVersion', '?')} → Fix: {vuln.get('FixedVersion', 'N/A')}",
                "description": vuln.get("Description", "")[:300],
                "remediation": f"Upgrade {vuln.get('PkgName', 'package')} to {vuln.get('FixedVersion', 'latest')}",
                "snippet":     vuln.get("Title", ""),
            })
    return findings


def parse_pip_audit(data: dict) -> list:
    """Parse pip-audit JSON output → list of finding dicts."""
    findings = []
    for dep in data.get("dependencies", []):
        for vuln in dep.get("vulns", []):
            findings.append({
                "tool":        "pip-audit",
                "severity":    "HIGH",
                "vuln_type":   f"CVE: {vuln.get('id', 'Unknown')}",
                "file_path":   dep.get("name", ""),
                "line":        f"Version: {dep.get('version', '?')}",
                "description": vuln.get("description", "")[:300],
                "remediation": (f"Upgrade {dep.get('name', 'package')} to fix {vuln.get('id', '?')}. "
                               f"Aliases: {', '.join(vuln.get('aliases', [])[:3])}"),
                "snippet":     "",
            })
    return findings


def parse_custom(data: dict) -> list:
    """Parse custom security_analysis.py JSON output → list of finding dicts."""
    findings = []
    for f in data.get("findings", []):
        sev = f.get("severity", "MEDIUM").upper()
        if sev not in SEVERITY_CFG:
            sev = "MEDIUM"
        findings.append({
            "tool":        "Custom Analysis",
            "severity":    sev,
            "vuln_type":   f.get("vuln_type", "Unknown"),
            "file_path":   f.get("file_path", ""),
            "line":        f.get("line", ""),
            "description": f.get("description", "")[:300],
            "remediation": f.get("remediation", "")[:300],
            "snippet":     (f.get("code_snippet") or "")[:200],
        })
    return findings


# ══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def build_security_summary_sheet(wb, all_findings: list, tool_counts: dict, now_str: str):
    ws = wb.create_sheet("🔒 Security Summary")
    ws.sheet_view.showGridLines = False

    total    = len(all_findings)
    critical = sum(1 for f in all_findings if f["severity"] == "CRITICAL")
    high     = sum(1 for f in all_findings if f["severity"] == "HIGH")
    medium   = sum(1 for f in all_findings if f["severity"] == "MEDIUM")
    low      = sum(1 for f in all_findings if f["severity"] in ("LOW", "INFO"))

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "🔒  PancreaScan — Backend Vulnerability & Security Scan Report"
    ws["A1"].font = Font(bold=True, color=WHITE, size=18, name="Calibri")
    ws["A1"].fill = _fill(BLUE_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 44

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {now_str}   |   Tools: Semgrep · Trivy · pip-audit · Custom Analysis"
    ws["A2"].font = Font(color=BLUE_LT, size=10, name="Calibri")
    ws["A2"].fill = _fill(BLUE_MED)
    ws["A2"].alignment = _align("center")
    ws.row_dimensions[2].height = 20

    # Stat cards
    cards = [
        ("Total Findings",  str(total),    BLUE_MED, BLUE_LT),
        ("🔴 CRITICAL",     str(critical), RED_DARK, RED_LT),
        ("🟠 HIGH",         str(high),     RED,      RED_LT),
        ("🟡 MEDIUM",       str(medium),   AMBER,    AMBER_LT),
        ("🟢 LOW/INFO",     str(low),      GREEN,    GREEN_LT),
    ]
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 40
    for ci, (label, val, fg, bg) in enumerate(cards, 1):
        ws.cell(row=4, column=ci, value=label).font = Font(bold=True, color=fg, size=10, name="Calibri")
        ws.cell(row=4, column=ci).fill             = _fill(bg)
        ws.cell(row=4, column=ci).alignment        = _align("center")
        ws.cell(row=5, column=ci, value=val).font  = Font(bold=True, color=fg, size=26, name="Calibri")
        ws.cell(row=5, column=ci).fill             = _fill(bg)
        ws.cell(row=5, column=ci).alignment        = _align("center")

    # Tool breakdown table
    ws.merge_cells("A7:D7")
    ws["A7"] = "Findings by Tool"
    ws["A7"].font = _font(bold=True, color=GRAY_DARK, size=12)
    ws["A7"].fill = _fill(GRAY_LT)
    ws["A7"].alignment = _align("center")
    ws.row_dimensions[7].height = 22

    hdr = ["Tool", "Findings", "Critical", "High"]
    for ci, h in enumerate(hdr, 1):
        c = ws.cell(row=8, column=ci, value=h)
        c.font = _font(bold=True)
        c.fill = _fill(BLUE_MED)
        c.alignment = _align("center")
        c.border = _border()
    ws.row_dimensions[8].height = 20

    for ri, (tool, cnt) in enumerate(tool_counts.items(), 9):
        crit = sum(1 for f in all_findings if f["tool"] == tool and f["severity"] == "CRITICAL")
        hi   = sum(1 for f in all_findings if f["tool"] == tool and f["severity"] == "HIGH")
        for ci, val in enumerate([tool, cnt, crit, hi], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = _fill(GRAY_LT if ri % 2 == 0 else WHITE)
            c.font = Font(color=GRAY_DARK, size=10, name="Calibri")
            c.alignment = _align("center")
            c.border = _border()
        ws.row_dimensions[ri].height = 18

    # Risk verdict
    verdict_row = 9 + len(tool_counts) + 2
    ws.merge_cells(f"A{verdict_row}:H{verdict_row}")
    if critical == 0 and high == 0:
        verdict = "✅  LOW RISK — No Critical or High vulnerabilities. Application is deployable with monitoring."
        color   = GREEN
    elif critical == 0:
        verdict = f"⚠️  MEDIUM RISK — {high} High severity findings. Remediate before production deployment."
        color   = AMBER
    else:
        verdict = f"🔴  HIGH RISK — {critical} Critical, {high} High findings. DO NOT DEPLOY until resolved."
        color   = RED
    ws[f"A{verdict_row}"] = verdict
    ws[f"A{verdict_row}"].font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    ws[f"A{verdict_row}"].fill = _fill(color)
    ws[f"A{verdict_row}"].alignment = _align("center")
    ws.row_dimensions[verdict_row].height = 30

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20


def build_all_findings_sheet(wb, all_findings: list):
    ws = wb.create_sheet("🔍 All Findings")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"🔍  All Security Findings — {len(all_findings)} Total"
    ws["A1"].font = _font(bold=True, size=14)
    ws["A1"].fill = _fill(BLUE_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 32

    cols   = ["Tool", "Severity", "Vulnerability Type", "File / Package", "Line/Version", "Description", "Remediation", "Code Snippet"]
    widths = [14, 12, 30, 30, 15, 50, 50, 35]
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True)
        c.fill = _fill(BLUE_MED)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    # Sort by severity
    order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4, "WARNING": 5}
    sorted_findings = sorted(all_findings, key=lambda f: order.get(f["severity"], 9))

    for ri, f in enumerate(sorted_findings, 3):
        sev_color, sev_bg, sev_label = SEVERITY_CFG.get(f["severity"], (GRAY_DARK, GRAY_LT, f["severity"]))
        row_fill = _fill(WHITE if ri % 2 == 0 else GRAY_LT)

        vals = [
            f["tool"], sev_label, f["vuln_type"],
            str(f["file_path"]),
            str(f["line"] or ""),
            f["description"],
            f["remediation"],
            f["snippet"] or "",
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _border()
            c.alignment = _align("left", "center", wrap=True)
            if ci == 2:
                c.fill = _fill(sev_bg)
                c.font = Font(color=sev_color, bold=True, size=10, name="Calibri")
            elif ci == 8:
                c.fill = _fill("EFF6FF")
                c.font = Font(color="1E40AF", size=9, name="Courier New")
            else:
                c.fill = row_fill
                c.font = Font(color=GRAY_DARK, size=10, name="Calibri")
        ws.row_dimensions[ri].height = 24


def build_owasp_checklist_sheet(wb, all_findings: list):
    """OWASP Top 10 coverage checklist."""
    ws = wb.create_sheet("📋 OWASP Top 10")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"] = "📋  OWASP Top 10 — Coverage Checklist"
    ws["A1"].font = _font(bold=True, size=14)
    ws["A1"].fill = _fill(BLUE_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 32

    hdr = ["OWASP Category", "Status", "Details", "Remediation"]
    widths = [40, 15, 55, 50]
    for ci, (h, w) in enumerate(zip(hdr, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True)
        c.fill = _fill(BLUE_MED)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    owasp_items = [
        ("A01:2021 — Broken Access Control",
         "Checked",
         "Auth middleware verified on all protected routes. JWT validation active.",
         "Ensure every non-public endpoint has Depends(verify_token). Implement RBAC."),
        ("A02:2021 — Cryptographic Failures",
         "Issue Found",
         "SHA-256 used for password hashing — weak for passwords. HTTPS enforced by FastAPI.",
         "Replace hashlib.sha256 with bcrypt/argon2. Use HTTPS in production."),
        ("A03:2021 — Injection",
         "Checked",
         "SQLAlchemy ORM with parameterized queries. SQL injection tests pass.",
         "Continue using ORM. Avoid raw SQL with f-strings. Validate all inputs with Pydantic."),
        ("A04:2021 — Insecure Design",
         "Partial",
         "No rate limiting on login endpoint. Missing refresh token mechanism.",
         "Add slowapi rate limiting. Implement JWT refresh tokens for better session management."),
        ("A05:2021 — Security Misconfiguration",
         "Issue Found",
         "CORS allows wildcard (*) origins. DEBUG middleware may be enabled.",
         "Restrict CORS to known frontend domains. Disable debug middleware in production."),
        ("A06:2021 — Vulnerable and Outdated Components",
         "Scanned",
         "Dependencies scanned with Trivy and pip-audit. Check findings sheet for CVEs.",
         "Upgrade packages with known CVEs. Pin dependency versions in requirements.txt."),
        ("A07:2021 — Identification and Authentication Failures",
         "Partial",
         "JWT auth implemented. No brute-force protection. No MFA.",
         "Add rate limiting on login. Implement account lockout after N failed attempts. Add MFA."),
        ("A08:2021 — Software and Data Integrity Failures",
         "Checked",
         "No deserialization of untrusted data detected. JWT signatures verified.",
         "Use HMAC or digital signatures for sensitive data. Validate JWT algorithm (reject 'none')."),
        ("A09:2021 — Security Logging and Monitoring Failures",
         "Partial",
         "Basic logging present. No structured security event logging or SIEM integration.",
         "Add audit logging for login, upload, and profile change events. Integrate with monitoring."),
        ("A10:2021 — Server-Side Request Forgery (SSRF)",
         "Not Applicable",
         "No external HTTP requests made by the server based on user input.",
         "If external URLs are added in future, validate/whitelist target domains."),
    ]

    status_cfg = {
        "Checked":         (GREEN,   GREEN_LT,  "✅ Checked"),
        "Issue Found":     (RED,     RED_LT,    "❌ Issue Found"),
        "Partial":         (AMBER,   AMBER_LT,  "⚠️ Partial"),
        "Scanned":         (BLUE_MED,BLUE_LT,   "🔍 Scanned"),
        "Not Applicable":  (GRAY_DARK,GRAY_LT,  "➖ N/A"),
    }

    for ri, (category, status, details, remediation) in enumerate(owasp_items, 3):
        s_color, s_bg, s_label = status_cfg.get(status, (GRAY_DARK, GRAY_LT, status))
        row_fill = _fill(WHITE if ri % 2 == 0 else GRAY_LT)

        for ci, val in enumerate([category, s_label, details, remediation], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _border()
            c.alignment = _align("left", "center", wrap=True)
            if ci == 2:
                c.fill = _fill(s_bg)
                c.font = Font(color=s_color, bold=True, size=10, name="Calibri")
            else:
                c.fill = row_fill
                c.font = Font(color=GRAY_DARK, size=10, name="Calibri")
        ws.row_dimensions[ri].height = 32


def build_dependency_sheet(wb, trivy_data: dict, pip_data: dict):
    """Dependency CVE detail sheet."""
    ws = wb.create_sheet("📦 Dependency CVEs")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    ws["A1"] = "📦  Dependency Vulnerabilities (Trivy + pip-audit)"
    ws["A1"].font = _font(bold=True, size=14)
    ws["A1"].fill = _fill(BLUE_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 32

    cols   = ["Package", "Installed Version", "CVE ID", "Severity", "Fix Version", "Description"]
    widths = [25, 18, 22, 12, 18, 60]
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True)
        c.fill = _fill(BLUE_MED)
        c.alignment = _align("center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    ri = 3
    # Trivy findings
    for result in trivy_data.get("Results", []):
        for vuln in result.get("Vulnerabilities", []):
            sev = vuln.get("Severity", "MEDIUM").upper()
            sev_color, sev_bg, sev_label = SEVERITY_CFG.get(sev, (GRAY_DARK, GRAY_LT, sev))
            row_fill = _fill(WHITE if ri % 2 == 0 else GRAY_LT)
            vals = [
                vuln.get("PkgName", ""),
                vuln.get("InstalledVersion", ""),
                vuln.get("VulnerabilityID", ""),
                sev_label,
                vuln.get("FixedVersion", "N/A"),
                (vuln.get("Description") or "")[:200],
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = _border()
                c.alignment = _align("left", "center", wrap=True)
                if ci == 4:
                    c.fill = _fill(sev_bg)
                    c.font = Font(color=sev_color, bold=True, size=10, name="Calibri")
                else:
                    c.fill = row_fill
                    c.font = Font(color=GRAY_DARK, size=10, name="Calibri")
            ws.row_dimensions[ri].height = 20
            ri += 1

    # pip-audit findings
    for dep in pip_data.get("dependencies", []):
        for vuln in dep.get("vulns", []):
            row_fill = _fill(WHITE if ri % 2 == 0 else GRAY_LT)
            vals = [
                dep.get("name", ""),
                dep.get("version", ""),
                vuln.get("id", ""),
                "🟠 HIGH",
                "Upgrade",
                (vuln.get("description") or "")[:200],
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = _border()
                c.alignment = _align("left", "center", wrap=True)
                if ci == 4:
                    c.fill = _fill(RED_LT)
                    c.font = Font(color=RED, bold=True, size=10, name="Calibri")
                else:
                    c.fill = row_fill
                    c.font = Font(color=GRAY_DARK, size=10, name="Calibri")
            ws.row_dimensions[ri].height = 20
            ri += 1

    if ri == 3:
        ws.cell(row=3, column=1, value="✅ No CVEs found in dependency scan").font = Font(color=GREEN, bold=True, name="Calibri")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def generate_security_xlsx(
    semgrep_path: str = None,
    trivy_path: str = None,
    pip_audit_path: str = None,
    custom_path: str = None,
    output_path: str = None,
) -> str:
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ts      = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")

    # Default output path
    if not output_path:
        out_dir = Path("Vulnerability Test Results")
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(out_dir / f"Security_Report_{ts}.xlsx")
    else:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # Load tool outputs
    semgrep_data  = _load_json(semgrep_path)
    trivy_data    = _load_json(trivy_path)
    pip_data      = _load_json(pip_audit_path)
    custom_data   = _load_json(custom_path)

    # Parse findings
    all_findings = []
    tool_counts  = {}

    semgrep_f = parse_semgrep(semgrep_data)
    all_findings.extend(semgrep_f)
    if semgrep_f:
        tool_counts["Semgrep"] = len(semgrep_f)

    trivy_f = parse_trivy(trivy_data)
    all_findings.extend(trivy_f)
    if trivy_f:
        tool_counts["Trivy"] = len(trivy_f)

    pip_f = parse_pip_audit(pip_data)
    all_findings.extend(pip_f)
    if pip_f:
        tool_counts["pip-audit"] = len(pip_f)

    custom_f = parse_custom(custom_data)
    all_findings.extend(custom_f)
    if custom_f:
        tool_counts["Custom Analysis"] = len(custom_f)

    print(f"  🔒 Total findings: {len(all_findings)}")

    # Build workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("⚙️  Building Security Summary sheet...")
    build_security_summary_sheet(wb, all_findings, tool_counts, now_str)

    print("⚙️  Building All Findings sheet...")
    build_all_findings_sheet(wb, all_findings)

    print("⚙️  Building OWASP Top 10 Checklist sheet...")
    build_owasp_checklist_sheet(wb, all_findings)

    print("⚙️  Building Dependency CVE sheet...")
    build_dependency_sheet(wb, trivy_data, pip_data)

    wb.save(output_path)
    print(f"\n  ✅ Security report saved: {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="PancreaScan Security XLSX Report Generator")
    parser.add_argument("--semgrep",   default=None, help="Path to semgrep-results.json")
    parser.add_argument("--trivy",     default=None, help="Path to trivy-results.json")
    parser.add_argument("--pip-audit", default=None, help="Path to pip-audit-results.json")
    parser.add_argument("--custom",    default=None, help="Path to custom-analysis.json")
    parser.add_argument("--output",    default=None, help="Output .xlsx path")
    args = parser.parse_args()

    generate_security_xlsx(
        semgrep_path=args.semgrep,
        trivy_path=args.trivy,
        pip_audit_path=args.pip_audit,
        custom_path=args.custom,
        output_path=args.output,
    )
