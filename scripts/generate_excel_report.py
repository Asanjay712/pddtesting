#!/usr/bin/env python3
"""
generate_excel_report.py
=========================
Reads Playwright JUnit XML results and generates a rich Excel workbook:
  - Sheet 1: Summary (totals, pass %, pie chart)
  - Sheet 2: Test Details (name, status, duration, error message)
  - Sheet 3: Screenshots list

Usage:
    python scripts/generate_excel_report.py \
        --junit e2e/test-results/junit.xml \
        --output "Test Results/Excel/Automation_Test_Report.xlsx" \
        --base-url "https://Asanjay712.github.io/pddtesting/"
"""

import argparse
import os
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
    import xml.etree.ElementTree as ET
except ImportError as e:
    print(f'[ERROR] Missing dependency: {e}')
    print('Install with: pip install openpyxl lxml')
    sys.exit(1)


# ── Colour palette ────────────────────────────────────────────────────────────
PASS_GREEN   = 'FF22C55E'   # Tailwind green-500
FAIL_RED     = 'FFEF4444'   # Tailwind red-500
SKIP_YELLOW  = 'FFF59E0B'   # Tailwind amber-400
HEADER_BLUE  = 'FF2563EB'   # Tailwind blue-600
HEADER_WHITE = 'FFFFFFFF'
ROW_LIGHT    = 'FFEFF6FF'   # Tailwind blue-50
ROW_WHITE    = 'FFFFFFFF'


def thin_border():
    side = Side(style='thin', color='FFE2E8F0')
    return Border(left=side, right=side, top=side, bottom=side)


def header_style(ws, cell, text):
    cell.value    = text
    cell.font     = Font(name='Calibri', bold=True, color=HEADER_WHITE, size=11)
    cell.fill     = PatternFill('solid', fgColor=HEADER_BLUE)
    cell.alignment= Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border   = thin_border()


# ── JUnit XML parser ──────────────────────────────────────────────────────────

def parse_junit(junit_path: str) -> dict:
    """Parse JUnit XML and return test result summary."""
    result = {
        'total': 0, 'passed': 0, 'failed': 0, 'skipped': 0, 'errors': 0,
        'duration': 0.0, 'tests': []
    }

    if not Path(junit_path).exists():
        print(f'[WARN] JUnit file not found: {junit_path} — using empty results')
        return result

    try:
        tree = ET.parse(junit_path)
        root = tree.getroot()

        # Handle both <testsuites> and <testsuite> root elements
        if root.tag == 'testsuites':
            suites = root.findall('testsuite')
        elif root.tag == 'testsuite':
            suites = [root]
        else:
            suites = root.findall('.//testsuite')

        for suite in suites:
            for tc in suite.findall('testcase'):
                name     = tc.get('name', 'Unnamed Test')
                classname= tc.get('classname', '')
                duration = float(tc.get('time', '0') or '0')

                failure  = tc.find('failure')
                error    = tc.find('error')
                skipped  = tc.find('skipped')

                if failure is not None:
                    status  = 'FAILED'
                    message = (failure.get('message', '') or failure.text or '')[:500]
                    result['failed'] += 1
                elif error is not None:
                    status  = 'ERROR'
                    message = (error.get('message', '') or error.text or '')[:500]
                    result['errors'] += 1
                elif skipped is not None:
                    status  = 'SKIPPED'
                    message = skipped.get('message', '')
                    result['skipped'] += 1
                else:
                    status  = 'PASSED'
                    message = ''
                    result['passed'] += 1

                result['total']    += 1
                result['duration'] += duration
                result['tests'].append({
                    'name':      name,
                    'classname': classname,
                    'status':    status,
                    'duration':  round(duration, 3),
                    'message':   message,
                })

    except ET.ParseError as e:
        print(f'[WARN] Could not parse JUnit XML: {e}')

    return result


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_summary_sheet(wb, data: dict, base_url: str, run_date: str):
    ws = wb.active
    ws.title = '📊 Summary'
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18

    # Title row
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value     = '🏥 Medical AI Platform — Automation Test Report'
    title_cell.font      = Font(name='Calibri', bold=True, size=16, color=HEADER_BLUE)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # Meta info
    meta = [
        ('Run Date',       run_date),
        ('Environment',    base_url),
        ('Framework',      'Playwright'),
        ('Report Version', '1.0'),
    ]
    for row_idx, (label, value) in enumerate(meta, start=2):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True, color='FF64748B')
        ws.cell(row=row_idx, column=2, value=value)

    # Spacer
    start_row = len(meta) + 3

    # Summary table header
    header_cell = ws.cell(row=start_row, column=1, value='Metric')
    header_style(ws, header_cell, 'Metric')
    value_cell  = ws.cell(row=start_row, column=2, value='Value')
    header_style(ws, value_cell, 'Value')

    total   = data['total'] or 1
    passed  = data['passed']
    failed  = data['failed'] + data['errors']
    skipped = data['skipped']
    pass_pct= round(passed / total * 100, 1) if total > 0 else 0
    duration= round(data['duration'], 2)

    summary_rows = [
        ('Total Tests',    data['total']),
        ('✅ Passed',      passed),
        ('❌ Failed',      failed),
        ('⏭️  Skipped',    skipped),
        ('Pass Rate (%)',  f'{pass_pct}%'),
        ('Duration (s)',   duration),
    ]

    for i, (label, value) in enumerate(summary_rows, start=start_row + 1):
        row_fill = ROW_LIGHT if i % 2 == 0 else ROW_WHITE
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        for cell in [lc, vc]:
            cell.fill      = PatternFill('solid', fgColor=row_fill)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = thin_border()
        lc.font = Font(bold=True)

        # Colour value cell by outcome
        if 'Passed' in label:
            vc.fill = PatternFill('solid', fgColor=PASS_GREEN)
            vc.font = Font(bold=True, color=HEADER_WHITE)
        elif 'Failed' in label and value > 0:
            vc.fill = PatternFill('solid', fgColor=FAIL_RED)
            vc.font = Font(bold=True, color=HEADER_WHITE)
        elif 'Pass Rate' in label:
            colour = PASS_GREEN if pass_pct >= 80 else FAIL_RED
            vc.fill = PatternFill('solid', fgColor=colour)
            vc.font = Font(bold=True, color=HEADER_WHITE)

    # Pie chart (pass/fail/skip)
    chart_row = start_row
    data_start = start_row + 2

    # Write chart data in hidden columns (E, F)
    ws['E1'] = 'Status'
    ws['F1'] = 'Count'
    ws['E2'] = 'Passed'
    ws['F2'] = passed
    ws['E3'] = 'Failed'
    ws['F3'] = failed
    ws['E4'] = 'Skipped'
    ws['F4'] = skipped

    pie = PieChart()
    pie.title = 'Test Results'
    pie.style = 10

    labels = Reference(ws, min_col=5, min_row=2, max_row=4)
    data_ref = Reference(ws, min_col=6, min_row=1, max_row=4)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels)

    # Colour slices
    slices = [
        DataPoint(idx=0, explosion=0),
        DataPoint(idx=1, explosion=5),
        DataPoint(idx=2, explosion=0),
    ]
    pie.series[0].dPt = slices

    pie.width  = 14
    pie.height = 10
    ws.add_chart(pie, f'D{start_row}')

    return ws


def build_details_sheet(wb, data: dict):
    ws = wb.create_sheet(title='📋 Test Details')

    # Column widths
    col_widths = [8, 50, 12, 12, 60]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ['#', 'Test Name', 'Status', 'Duration (s)', 'Error / Notes']
    for col, h in enumerate(headers, start=1):
        header_style(ws, ws.cell(row=1, column=col), h)

    status_colours = {
        'PASSED':  PASS_GREEN,
        'FAILED':  FAIL_RED,
        'ERROR':   FAIL_RED,
        'SKIPPED': SKIP_YELLOW,
    }

    for row_idx, test in enumerate(data['tests'], start=2):
        row_fill = ROW_LIGHT if row_idx % 2 == 0 else ROW_WHITE
        status   = test['status']
        cells    = [
            (1, row_idx, str(row_idx - 1)),
            (2, row_idx, test['name']),
            (3, row_idx, status),
            (4, row_idx, test['duration']),
            (5, row_idx, test['message']),
        ]
        for col, r, val in cells:
            c            = ws.cell(row=r, column=col, value=val)
            c.fill       = PatternFill('solid', fgColor=row_fill)
            c.alignment  = Alignment(vertical='center', wrap_text=(col == 5))
            c.border     = thin_border()
            if col == 3:
                c.fill = PatternFill('solid', fgColor=status_colours.get(status, SKIP_YELLOW))
                c.font = Font(bold=True, color=HEADER_WHITE)
                c.alignment = Alignment(horizontal='center', vertical='center')

    # Freeze header row
    ws.freeze_panes = 'A2'
    return ws


def build_screenshots_sheet(wb, screenshots_dir: str):
    ws = wb.create_sheet(title='📸 Screenshots')
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 20

    header_style(ws, ws.cell(row=1, column=1), 'Screenshot File')
    header_style(ws, ws.cell(row=1, column=2), 'Size')

    screenshots = []
    if Path(screenshots_dir).exists():
        for f in sorted(Path(screenshots_dir).rglob('*.png')):
            screenshots.append((str(f), f'{f.stat().st_size // 1024} KB'))

    if not screenshots:
        ws.cell(row=2, column=1, value='No screenshots captured in this run').font = Font(italic=True, color='FF94A3B8')
        return ws

    for i, (path, size) in enumerate(screenshots, start=2):
        ws.cell(row=i, column=1, value=path).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=2, value=size)

    return ws


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--junit',            default='e2e/test-results/junit.xml')
    parser.add_argument('--output',           default='Test Results/Excel/Automation_Test_Report.xlsx')
    parser.add_argument('--base-url',         default='https://Asanjay712.github.io/pddtesting/')
    parser.add_argument('--screenshots-dir',  default='e2e/test-results')
    args = parser.parse_args()

    run_date = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    print(f'[INFO] Parsing JUnit XML: {args.junit}')
    data = parse_junit(args.junit)

    print(f'[INFO] Total={data["total"]} Passed={data["passed"]} Failed={data["failed"]} Skipped={data["skipped"]}')

    wb = openpyxl.Workbook()
    build_summary_sheet(wb, data, args.base_url, run_date)
    build_details_sheet(wb, data)
    build_screenshots_sheet(wb, args.screenshots_dir)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f'[INFO] Excel report saved: {out_path}')


if __name__ == '__main__':
    main()
